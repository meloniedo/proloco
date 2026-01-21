#!/usr/bin/env python3
"""
Script di backup automatico su USB
Si attiva quando viene inserita una chiavetta USB

Crea la struttura:
/USB/BarManager_Backup/
├── report_generale.xlsx
├── storico_completo.xlsx
└── mensili/
    ├── 2026-01/
    │   └── report_gennaio_2026.xlsx
    ├── 2026-02/
    │   └── report_febbraio_2026.xlsx
    └── ...
"""

import os
import sys
import shutil
import time
from datetime import datetime, timedelta
from pathlib import Path

# Aggiungi la directory corrente al path per importare i moduli
sys.path.insert(0, '/home/pi/bar_manager')

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("Errore: openpyxl non installato")
    sys.exit(1)

# Directory dati
DATA_DIR = Path('/home/pi/bar_manager/dati')
EXCEL_FILE = DATA_DIR / 'storico_bar.xlsx'

def trova_usb():
    """Trova il punto di mount della chiavetta USB"""
    usb_paths = [
        '/media/pi',
        '/mnt/usb',
        '/media'
    ]
    
    for base_path in usb_paths:
        if os.path.exists(base_path):
            for item in os.listdir(base_path):
                full_path = os.path.join(base_path, item)
                if os.path.ismount(full_path) or os.path.isdir(full_path):
                    # Verifica che sia scrivibile
                    try:
                        test_file = os.path.join(full_path, '.test_write')
                        with open(test_file, 'w') as f:
                            f.write('test')
                        os.remove(test_file)
                        return full_path
                    except:
                        continue
    return None

def get_vendite_spese():
    """Legge vendite e spese dal file Excel"""
    if not EXCEL_FILE.exists():
        return [], []
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    
    vendite = []
    ws_vendite = wb["Vendite"]
    for row in ws_vendite.iter_rows(min_row=2, values_only=True):
        if row[0]:
            vendite.append({
                "id": row[0],
                "data": row[1],
                "ora": row[2],
                "prodotto": row[3],
                "categoria": row[4],
                "importo": float(row[5]) if row[5] else 0,
                "timestamp": row[6]
            })
    
    spese = []
    ws_spese = wb["Spese"]
    for row in ws_spese.iter_rows(min_row=2, values_only=True):
        if row[0]:
            spese.append({
                "id": row[0],
                "data": row[1],
                "ora": row[2],
                "categoria": row[3],
                "note": row[4],
                "importo": float(row[5]) if row[5] else 0,
                "timestamp": row[6]
            })
    
    return vendite, spese

def crea_report_mese(vendite, spese, anno, mese, output_path):
    """Crea report per un mese specifico"""
    # Filtra per mese
    vendite_mese = [v for v in vendite if v["timestamp"] and 
                   v["timestamp"].startswith(f"{anno}-{mese:02d}")]
    spese_mese = [s for s in spese if s["timestamp"] and 
                 s["timestamp"].startswith(f"{anno}-{mese:02d}")]
    
    if not vendite_mese and not spese_mese:
        return False
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Report {mese:02d}/{anno}"
    
    # Stili
    header_fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Titolo
    nome_mese = datetime(anno, mese, 1).strftime("%B %Y")
    ws.append([f"REPORT {nome_mese.upper()}"])
    ws.append([])
    
    # Riepilogo
    totale_incasso = sum(v["importo"] for v in vendite_mese)
    totale_spese = sum(s["importo"] for s in spese_mese)
    profitto = totale_incasso - totale_spese
    
    ws.append(["RIEPILOGO"])
    ws.append(["Vendite totali:", len(vendite_mese)])
    ws.append(["Incasso totale:", f"€ {totale_incasso:.2f}"])
    ws.append(["Spese totali:", f"€ {totale_spese:.2f}"])
    ws.append(["PROFITTO NETTO:", f"€ {profitto:.2f}"])
    ws.append([])
    
    # Dettaglio vendite
    ws.append(["DETTAGLIO VENDITE"])
    headers = ["Data", "Ora", "Prodotto", "Categoria", "Importo €"]
    ws.append(headers)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
    
    for v in vendite_mese:
        ws.append([v["data"], v["ora"], v["prodotto"], v["categoria"], v["importo"]])
    
    ws.append([])
    
    # Dettaglio spese
    ws.append(["DETTAGLIO SPESE"])
    headers_spese = ["Data", "Ora", "Categoria", "Note", "Importo €"]
    ws.append(headers_spese)
    for col, header in enumerate(headers_spese, 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
    
    for s in spese_mese:
        ws.append([s["data"], s["ora"], s["categoria"], s["note"], s["importo"]])
    
    # Salva
    wb.save(output_path)
    return True

def crea_report_generale(vendite, spese, output_path):
    """Crea report generale con tutte le statistiche"""
    wb = openpyxl.Workbook()
    
    # Foglio riepilogo
    ws = wb.active
    ws.title = "Riepilogo Generale"
    
    totale_incasso = sum(v["importo"] for v in vendite)
    totale_spese = sum(s["importo"] for s in spese)
    profitto = totale_incasso - totale_spese
    
    ws.append(["REPORT GENERALE - PROLOCO SANTA BIANCA"])
    ws.append([f"Generato il: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
    ws.append([])
    ws.append(["STATISTICHE TOTALI"])
    ws.append(["Numero vendite:", len(vendite)])
    ws.append(["Incasso totale:", f"€ {totale_incasso:.2f}"])
    ws.append(["Spese totali:", f"€ {totale_spese:.2f}"])
    ws.append(["PROFITTO NETTO:", f"€ {profitto:.2f}"])
    
    # Top prodotti
    ws.append([])
    ws.append(["TOP 10 PRODOTTI PIÙ VENDUTI"])
    prodotti_count = {}
    for v in vendite:
        nome = v["prodotto"]
        if nome not in prodotti_count:
            prodotti_count[nome] = {"quantita": 0, "incasso": 0}
        prodotti_count[nome]["quantita"] += 1
        prodotti_count[nome]["incasso"] += v["importo"]
    
    top_prodotti = sorted(prodotti_count.items(), key=lambda x: x[1]["quantita"], reverse=True)[:10]
    ws.append(["Prodotto", "Quantità", "Incasso"])
    for nome, data in top_prodotti:
        ws.append([nome, data["quantita"], f"€ {data['incasso']:.2f}"])
    
    # Riepilogo per mese
    ws.append([])
    ws.append(["RIEPILOGO PER MESE"])
    ws.append(["Mese", "Vendite", "Incasso", "Spese", "Profitto"])
    
    mesi_vendite = {}
    for v in vendite:
        if v["timestamp"]:
            mese_key = v["timestamp"][:7]  # YYYY-MM
            if mese_key not in mesi_vendite:
                mesi_vendite[mese_key] = {"vendite": 0, "incasso": 0}
            mesi_vendite[mese_key]["vendite"] += 1
            mesi_vendite[mese_key]["incasso"] += v["importo"]
    
    mesi_spese = {}
    for s in spese:
        if s["timestamp"]:
            mese_key = s["timestamp"][:7]
            if mese_key not in mesi_spese:
                mesi_spese[mese_key] = 0
            mesi_spese[mese_key] += s["importo"]
    
    tutti_mesi = sorted(set(list(mesi_vendite.keys()) + list(mesi_spese.keys())))
    for mese in tutti_mesi:
        v_data = mesi_vendite.get(mese, {"vendite": 0, "incasso": 0})
        s_tot = mesi_spese.get(mese, 0)
        profitto_mese = v_data["incasso"] - s_tot
        ws.append([mese, v_data["vendite"], f"€ {v_data['incasso']:.2f}", 
                  f"€ {s_tot:.2f}", f"€ {profitto_mese:.2f}"])
    
    wb.save(output_path)

def esegui_backup(usb_path):
    """Esegue il backup completo sulla chiavetta USB"""
    print(f"🔄 Avvio backup su {usb_path}")
    
    # Crea cartella principale
    backup_dir = Path(usb_path) / "BarManager_Backup"
    backup_dir.mkdir(exist_ok=True)
    
    # Crea cartella mensili
    mensili_dir = backup_dir / "mensili"
    mensili_dir.mkdir(exist_ok=True)
    
    # Leggi dati
    vendite, spese = get_vendite_spese()
    
    if not vendite and not spese:
        print("⚠️ Nessun dato da backuppare")
        # Crea file vuoto di conferma
        with open(backup_dir / "backup_vuoto.txt", 'w') as f:
            f.write(f"Backup eseguito il {datetime.now().strftime('%d/%m/%Y %H:%M')}\n")
            f.write("Nessun dato presente nel database.\n")
        return
    
    # Copia file storico completo
    if EXCEL_FILE.exists():
        shutil.copy(EXCEL_FILE, backup_dir / "storico_completo.xlsx")
        print("✅ Copiato storico_completo.xlsx")
    
    # Crea report generale
    crea_report_generale(vendite, spese, backup_dir / "report_generale.xlsx")
    print("✅ Creato report_generale.xlsx")
    
    # Trova tutti i mesi con dati
    mesi_con_dati = set()
    for v in vendite:
        if v["timestamp"]:
            try:
                dt = datetime.fromisoformat(v["timestamp"].replace("Z", ""))
                mesi_con_dati.add((dt.year, dt.month))
            except:
                pass
    
    for s in spese:
        if s["timestamp"]:
            try:
                dt = datetime.fromisoformat(s["timestamp"].replace("Z", ""))
                mesi_con_dati.add((dt.year, dt.month))
            except:
                pass
    
    # Crea report per ogni mese
    for anno, mese in sorted(mesi_con_dati):
        mese_dir = mensili_dir / f"{anno}-{mese:02d}"
        mese_dir.mkdir(exist_ok=True)
        
        nome_mese = datetime(anno, mese, 1).strftime("%B").lower()
        report_path = mese_dir / f"report_{nome_mese}_{anno}.xlsx"
        
        if crea_report_mese(vendite, spese, anno, mese, report_path):
            print(f"✅ Creato report {nome_mese} {anno}")
    
    # File di conferma backup
    with open(backup_dir / "ultimo_backup.txt", 'w') as f:
        f.write(f"Backup completato: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n")
        f.write(f"Vendite totali: {len(vendite)}\n")
        f.write(f"Spese totali: {len(spese)}\n")
    
    print(f"✅ Backup completato in {backup_dir}")

def main():
    """Funzione principale - chiamata da udev"""
    print(f"\n{'='*50}")
    print(f"📀 BACKUP USB - {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"{'='*50}")
    
    # Aspetta che la USB sia montata
    time.sleep(3)
    
    # Trova USB
    usb_path = trova_usb()
    
    if not usb_path:
        print("❌ Nessuna chiavetta USB trovata")
        return
    
    print(f"📁 USB trovata: {usb_path}")
    
    try:
        esegui_backup(usb_path)
    except Exception as e:
        print(f"❌ Errore backup: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
