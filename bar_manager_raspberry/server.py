#!/usr/bin/env python3
"""
Bar Manager - Server per Raspberry Pi
Proloco Santa Bianca

Salva i dati in Excel e gestisce report automatici via email.
"""

import os
import json
import smtplib
import socket
import threading
import time
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path
from typing import Optional
import uuid

from fastapi import FastAPI, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import uvicorn
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Import configurazione
from config import CONFIG

# ==================== SETUP ====================

app = FastAPI(title="Bar Manager API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Directory dati
DATA_DIR = Path(__file__).parent / "dati"
DATA_DIR.mkdir(exist_ok=True)

EXCEL_FILE = DATA_DIR / "storico_bar.xlsx"
LAST_REPORT_FILE = DATA_DIR / "ultimo_report.txt"

# ==================== MODELLI ====================

class Vendita(BaseModel):
    id: Optional[str] = None
    prodotto_id: str
    nome_prodotto: str
    prezzo: float
    categoria: str
    timestamp: Optional[str] = None

class Spesa(BaseModel):
    id: Optional[str] = None
    categoria_spesa: str
    importo: float
    note: Optional[str] = ""
    timestamp: Optional[str] = None

class ResetRequest(BaseModel):
    password: str
    periodo: str

class DownloadRequest(BaseModel):
    password: str

# ==================== GESTIONE REPORT SETTIMANALE ====================

REPORT_WEEK_FILE = DATA_DIR / "report_settimana_corrente.txt"

def get_lunedi_corrente():
    """Ritorna il lunedì della settimana corrente"""
    now = datetime.now()
    lunedi = now - timedelta(days=now.weekday())
    return lunedi.replace(hour=0, minute=0, second=0, microsecond=0)

def is_report_inviato_questa_settimana():
    """Controlla se il report è stato già inviato questa settimana"""
    if not REPORT_WEEK_FILE.exists():
        return False
    
    try:
        with open(REPORT_WEEK_FILE) as f:
            data_str = f.read().strip()
            data_ultimo = datetime.fromisoformat(data_str)
            lunedi = get_lunedi_corrente()
            return data_ultimo >= lunedi
    except:
        return False

def segna_report_inviato():
    """Segna che il report è stato inviato questa settimana"""
    with open(REPORT_WEEK_FILE, 'w') as f:
        f.write(datetime.now().isoformat())

# ==================== GESTIONE EXCEL ====================

def init_excel():
    """Inizializza il file Excel se non esiste"""
    if EXCEL_FILE.exists():
        return
    
    wb = openpyxl.Workbook()
    
    # Foglio Vendite
    ws_vendite = wb.active
    ws_vendite.title = "Vendite"
    headers_vendite = ["ID", "Data", "Ora", "Prodotto", "Categoria", "Importo €", "Timestamp"]
    ws_vendite.append(headers_vendite)
    style_headers(ws_vendite, len(headers_vendite))
    
    # Foglio Spese
    ws_spese = wb.create_sheet("Spese")
    headers_spese = ["ID", "Data", "Ora", "Categoria", "Note", "Importo €", "Timestamp"]
    ws_spese.append(headers_spese)
    style_headers(ws_spese, len(headers_spese))
    
    # Foglio Prodotti
    ws_prodotti = wb.create_sheet("Prodotti")
    headers_prodotti = ["ID", "Nome", "Prezzo €", "Categoria", "Icona"]
    ws_prodotti.append(headers_prodotti)
    style_headers(ws_prodotti, len(headers_prodotti))
    
    # Popola prodotti da config
    for i, prod in enumerate(CONFIG["listino"]):
        ws_prodotti.append([
            f"prod_{i+1}",
            prod["nome"],
            prod["prezzo"],
            prod["categoria"],
            prod["icona"]
        ])
    
    wb.save(EXCEL_FILE)
    print(f"✅ File Excel creato: {EXCEL_FILE}")

def style_headers(ws, num_cols):
    """Applica stile alle intestazioni"""
    header_fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

def get_vendite():
    """Legge tutte le vendite dal file Excel"""
    if not EXCEL_FILE.exists():
        init_excel()
        return []
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Vendite"]
    
    vendite = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Se ha ID
            vendite.append({
                "id": row[0],
                "timestamp": row[6] if row[6] else f"{row[1]}T{row[2]}",
                "nome_prodotto": row[3],
                "categoria": row[4],
                "prezzo": float(row[5]) if row[5] else 0,
                "prodotto_id": row[0]
            })
    
    # Ordina per timestamp decrescente
    vendite.sort(key=lambda x: x["timestamp"], reverse=True)
    return vendite

def get_spese():
    """Legge tutte le spese dal file Excel"""
    if not EXCEL_FILE.exists():
        init_excel()
        return []
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Spese"]
    
    spese = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Se ha ID
            spese.append({
                "id": row[0],
                "timestamp": row[6] if row[6] else f"{row[1]}T{row[2]}",
                "categoria_spesa": row[3],
                "note": row[4] or "",
                "importo": float(row[5]) if row[5] else 0
            })
    
    spese.sort(key=lambda x: x["timestamp"], reverse=True)
    return spese

def get_prodotti():
    """Legge i prodotti dal file Excel"""
    if not EXCEL_FILE.exists():
        init_excel()
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Prodotti"]
    
    prodotti = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            prodotti.append({
                "id": row[0],
                "nome": row[1],
                "prezzo": float(row[2]) if row[2] else 0,
                "categoria": row[3],
                "icona": row[4]
            })
    
    return prodotti

def add_vendita(vendita: Vendita):
    """Aggiunge una vendita al file Excel"""
    if not EXCEL_FILE.exists():
        init_excel()
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Vendite"]
    
    now = datetime.now()
    vendita_id = vendita.id or str(uuid.uuid4())[:8]
    timestamp = vendita.timestamp or now.isoformat()
    
    ws.append([
        vendita_id,
        now.strftime("%d/%m/%Y"),
        now.strftime("%H:%M:%S"),
        vendita.nome_prodotto,
        vendita.categoria,
        vendita.prezzo,
        timestamp
    ])
    
    wb.save(EXCEL_FILE)
    return {"id": vendita_id, "timestamp": timestamp}

def add_spesa(spesa: Spesa):
    """Aggiunge una spesa al file Excel"""
    if not EXCEL_FILE.exists():
        init_excel()
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Spese"]
    
    now = datetime.now()
    spesa_id = spesa.id or str(uuid.uuid4())[:8]
    timestamp = spesa.timestamp or now.isoformat()
    
    ws.append([
        spesa_id,
        now.strftime("%d/%m/%Y"),
        now.strftime("%H:%M:%S"),
        spesa.categoria_spesa,
        spesa.note or "",
        spesa.importo,
        timestamp
    ])
    
    wb.save(EXCEL_FILE)
    return {"id": spesa_id, "timestamp": timestamp}

def delete_vendita(vendita_id: str):
    """Elimina una vendita dal file Excel"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Vendite"]
    
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == vendita_id:
            ws.delete_rows(row)
            break
    
    wb.save(EXCEL_FILE)

def delete_spesa(spesa_id: str):
    """Elimina una spesa dal file Excel"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Spese"]
    
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == spesa_id:
            ws.delete_rows(row)
            break
    
    wb.save(EXCEL_FILE)

def reset_periodo(periodo: str):
    """Resetta vendite e spese per un periodo"""
    now = datetime.now()
    
    if periodo == "oggi":
        data_inizio = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif periodo == "settimana":
        data_inizio = now - timedelta(days=7)
    else:  # mese
        data_inizio = now - timedelta(days=30)
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    
    # Reset vendite
    ws_vendite = wb["Vendite"]
    rows_to_delete = []
    for row in range(2, ws_vendite.max_row + 1):
        timestamp = ws_vendite.cell(row=row, column=7).value
        if timestamp:
            try:
                dt = datetime.fromisoformat(timestamp.replace("Z", ""))
                if dt >= data_inizio:
                    rows_to_delete.append(row)
            except:
                pass
    
    for row in reversed(rows_to_delete):
        ws_vendite.delete_rows(row)
    
    # Reset spese
    ws_spese = wb["Spese"]
    rows_to_delete = []
    for row in range(2, ws_spese.max_row + 1):
        timestamp = ws_spese.cell(row=row, column=7).value
        if timestamp:
            try:
                dt = datetime.fromisoformat(timestamp.replace("Z", ""))
                if dt >= data_inizio:
                    rows_to_delete.append(row)
            except:
                pass
    
    for row in reversed(rows_to_delete):
        ws_spese.delete_rows(row)
    
    wb.save(EXCEL_FILE)

# ==================== STATISTICHE ====================

def calcola_statistiche(data_inizio: datetime, data_fine: datetime):
    """Calcola statistiche per un periodo"""
    vendite = get_vendite()
    spese = get_spese()
    
    # Filtra per periodo
    vendite_periodo = []
    for v in vendite:
        try:
            dt = datetime.fromisoformat(v["timestamp"].replace("Z", ""))
            if data_inizio <= dt <= data_fine:
                vendite_periodo.append(v)
        except:
            pass
    
    spese_periodo = []
    for s in spese:
        try:
            dt = datetime.fromisoformat(s["timestamp"].replace("Z", ""))
            if data_inizio <= dt <= data_fine:
                spese_periodo.append(s)
        except:
            pass
    
    totale_vendite = len(vendite_periodo)
    totale_incasso = sum(v["prezzo"] for v in vendite_periodo)
    totale_spese = sum(s["importo"] for s in spese_periodo)
    profitto_netto = totale_incasso - totale_spese
    
    # Per categoria
    per_categoria = {}
    for v in vendite_periodo:
        cat = v["categoria"]
        if cat not in per_categoria:
            per_categoria[cat] = {"vendite": 0, "incasso": 0}
        per_categoria[cat]["vendite"] += 1
        per_categoria[cat]["incasso"] += v["prezzo"]
    
    # Top prodotti
    prodotti_count = {}
    for v in vendite_periodo:
        nome = v["nome_prodotto"]
        if nome not in prodotti_count:
            prodotti_count[nome] = {"nome": nome, "quantita": 0, "incasso": 0}
        prodotti_count[nome]["quantita"] += 1
        prodotti_count[nome]["incasso"] += v["prezzo"]
    
    top_prodotti = sorted(prodotti_count.values(), key=lambda x: x["quantita"], reverse=True)[:5]
    
    return {
        "totale_vendite": totale_vendite,
        "totale_incasso": totale_incasso,
        "totale_spese": totale_spese,
        "profitto_netto": profitto_netto,
        "per_categoria": per_categoria,
        "top_prodotti": top_prodotti
    }

# ==================== REPORT EMAIL ====================

def genera_report_excel():
    """Genera file Excel con report completo per email"""
    report_file = DATA_DIR / f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    wb = openpyxl.Workbook()
    now = datetime.now()
    
    # Calcola periodi
    # Lunedì di questa settimana
    lunedi_corrente = now - timedelta(days=now.weekday())
    lunedi_corrente = lunedi_corrente.replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Lunedì settimana scorsa
    lunedi_scorso = lunedi_corrente - timedelta(days=7)
    domenica_scorsa = lunedi_corrente - timedelta(seconds=1)
    
    # Inizio mese
    inizio_mese = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Data installazione (prima vendita o oggi)
    vendite = get_vendite()
    if vendite:
        prima_vendita = min(datetime.fromisoformat(v["timestamp"].replace("Z", "")) for v in vendite)
    else:
        prima_vendita = now
    
    # FOGLIO 1: Settimana in corso
    ws1 = wb.active
    ws1.title = "Settimana Corrente"
    stats1 = calcola_statistiche(lunedi_corrente, now)
    scrivi_foglio_report(ws1, f"SETTIMANA IN CORSO ({lunedi_corrente.strftime('%d/%m')} - {now.strftime('%d/%m/%Y')})", stats1)
    
    # FOGLIO 2: Settimana scorsa
    ws2 = wb.create_sheet("Settimana Scorsa")
    stats2 = calcola_statistiche(lunedi_scorso, domenica_scorsa)
    scrivi_foglio_report(ws2, f"SETTIMANA SCORSA ({lunedi_scorso.strftime('%d/%m')} - {domenica_scorsa.strftime('%d/%m/%Y')})", stats2)
    
    # FOGLIO 3: Mese in corso
    ws3 = wb.create_sheet("Mese Corrente")
    stats3 = calcola_statistiche(inizio_mese, now)
    scrivi_foglio_report(ws3, f"MESE DI {now.strftime('%B %Y').upper()}", stats3)
    
    # FOGLIO 4: Report totale
    ws4 = wb.create_sheet("Totale Generale")
    stats4 = calcola_statistiche(prima_vendita, now)
    scrivi_foglio_report(ws4, f"TOTALE DALL'INSTALLAZIONE ({prima_vendita.strftime('%d/%m/%Y')} - {now.strftime('%d/%m/%Y')})", stats4)
    
    # FOGLIO 5: Dettaglio vendite settimana
    ws5 = wb.create_sheet("Dettaglio Vendite")
    ws5.append(["DETTAGLIO VENDITE SETTIMANA IN CORSO"])
    ws5.append([])
    ws5.append(["Data", "Ora", "Prodotto", "Categoria", "Importo €"])
    
    for v in vendite:
        try:
            dt = datetime.fromisoformat(v["timestamp"].replace("Z", ""))
            if dt >= lunedi_corrente:
                ws5.append([
                    dt.strftime("%d/%m/%Y"),
                    dt.strftime("%H:%M"),
                    v["nome_prodotto"],
                    v["categoria"],
                    v["prezzo"]
                ])
        except:
            pass
    
    wb.save(report_file)
    return report_file

def scrivi_foglio_report(ws, titolo, stats):
    """Scrive le statistiche in un foglio Excel"""
    # Stili
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    money_font = Font(bold=True, size=12)
    
    row = 1
    ws.cell(row=row, column=1, value=titolo).font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    
    row += 2
    ws.cell(row=row, column=1, value="RIEPILOGO").font = Font(bold=True)
    
    row += 1
    ws.cell(row=row, column=1, value="Numero Vendite:")
    ws.cell(row=row, column=2, value=stats["totale_vendite"])
    
    row += 1
    ws.cell(row=row, column=1, value="Incasso Totale:")
    ws.cell(row=row, column=2, value=f"€ {stats['totale_incasso']:.2f}").font = money_font
    
    row += 1
    ws.cell(row=row, column=1, value="Spese Totali:")
    ws.cell(row=row, column=2, value=f"€ {stats['totale_spese']:.2f}")
    
    row += 1
    ws.cell(row=row, column=1, value="PROFITTO NETTO:")
    profit_cell = ws.cell(row=row, column=2, value=f"€ {stats['profitto_netto']:.2f}")
    profit_cell.font = Font(bold=True, size=14, color="228B22" if stats['profitto_netto'] >= 0 else "FF0000")
    
    row += 2
    ws.cell(row=row, column=1, value="PER CATEGORIA").font = Font(bold=True)
    
    row += 1
    for cat, data in stats["per_categoria"].items():
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=f"{data['vendite']} vendite")
        ws.cell(row=row, column=3, value=f"€ {data['incasso']:.2f}")
        row += 1
    
    row += 1
    ws.cell(row=row, column=1, value="TOP 5 PRODOTTI").font = Font(bold=True)
    
    row += 1
    for i, prod in enumerate(stats["top_prodotti"], 1):
        ws.cell(row=row, column=1, value=f"{i}. {prod['nome']}")
        ws.cell(row=row, column=2, value=f"{prod['quantita']} vendite")
        ws.cell(row=row, column=3, value=f"€ {prod['incasso']:.2f}")
        row += 1
    
    # Larghezza colonne
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

def invia_email_report():
    """Invia il report via email"""
    try:
        # Genera report Excel
        report_file = genera_report_excel()
        
        now = datetime.now()
        
        # Calcola statistiche per email
        lunedi_corrente = now - timedelta(days=now.weekday())
        lunedi_corrente = lunedi_corrente.replace(hour=0, minute=0, second=0, microsecond=0)
        stats = calcola_statistiche(lunedi_corrente, now)
        
        # Prepara email
        msg = MIMEMultipart()
        msg['From'] = CONFIG["email_mittente"]
        msg['To'] = ", ".join(CONFIG["email_destinatari"])
        msg['Subject'] = f"📊 Report {CONFIG['nome_bar']} - {now.strftime('%d/%m/%Y %H:%M')}"
        
        # Corpo email
        body = f"""
Buongiorno!

Ecco il report automatico di {CONFIG['nome_bar']}.

═══════════════════════════════════════
📅 SETTIMANA IN CORSO (da Lunedì)
═══════════════════════════════════════
• Vendite: {stats['totale_vendite']}
• Incasso: € {stats['totale_incasso']:.2f}
• Spese: € {stats['totale_spese']:.2f}
• PROFITTO: € {stats['profitto_netto']:.2f}

═══════════════════════════════════════
🏆 TOP 3 PRODOTTI
═══════════════════════════════════════
"""
        for i, prod in enumerate(stats["top_prodotti"][:3], 1):
            body += f"{i}. {prod['nome']}: {prod['quantita']} vendite (€ {prod['incasso']:.2f})\n"
        
        body += f"""
═══════════════════════════════════════

📎 In allegato trovi il report Excel completo con:
   • Settimana in corso
   • Settimana scorsa
   • Mese corrente
   • Totale generale

Generato automaticamente da Bar Manager
{now.strftime('%d/%m/%Y alle %H:%M')}
"""
        
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        # Allega Excel
        with open(report_file, 'rb') as f:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="report_{now.strftime("%Y%m%d")}.xlsx"')
            msg.attach(part)
        
        # Invia
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(CONFIG["email_mittente"], CONFIG["email_password"])
        server.send_message(msg)
        server.quit()
        
        # Salva timestamp ultimo report
        with open(LAST_REPORT_FILE, 'w') as f:
            f.write(now.isoformat())
        
        # Pulisci file temporaneo
        report_file.unlink()
        
        print(f"✅ Report inviato a: {', '.join(CONFIG['email_destinatari'])}")
        return True
        
    except Exception as e:
        print(f"❌ Errore invio email: {e}")
        return False

def check_internet():
    """Verifica se c'è connessione internet"""
    try:
        socket.create_connection(("8.8.8.8", 53), timeout=3)
        return True
    except OSError:
        return False

def get_ultimo_report():
    """Ritorna la data dell'ultimo report inviato"""
    if LAST_REPORT_FILE.exists():
        with open(LAST_REPORT_FILE) as f:
            return datetime.fromisoformat(f.read().strip())
    return None

def monitor_internet():
    """Thread che monitora la connessione internet e invia report"""
    print("🔍 Monitoraggio connessione internet avviato...")
    
    was_offline = True
    
    while True:
        try:
            is_online = check_internet()
            
            if is_online and was_offline:
                print("📶 Connessione internet rilevata!")
                
                # Controlla se dobbiamo inviare report
                ultimo = get_ultimo_report()
                now = datetime.now()
                
                if ultimo is None:
                    # Mai inviato, invia subito
                    print("📧 Primo report, invio...")
                    invia_email_report()
                else:
                    ore_passate = (now - ultimo).total_seconds() / 3600
                    if ore_passate >= CONFIG["ore_minime_tra_report"]:
                        print(f"📧 Sono passate {ore_passate:.1f} ore dall'ultimo report, invio...")
                        invia_email_report()
                    else:
                        print(f"⏳ Ultimo report {ore_passate:.1f} ore fa, attendo...")
            
            was_offline = not is_online
            
        except Exception as e:
            print(f"⚠️ Errore monitor: {e}")
        
        # Controlla ogni 30 secondi
        time.sleep(30)

# ==================== API ROUTES ====================

@app.get("/api/config")
async def api_get_config():
    """Ritorna la configurazione"""
    return {
        "nome_bar": CONFIG["nome_bar"],
        "listino": CONFIG["listino"],
        "categorie_spese": CONFIG["categorie_spese"],
        "password_reset": CONFIG["password_reset"]
    }

@app.get("/api/prodotti")
async def api_get_prodotti():
    """Ritorna tutti i prodotti"""
    return get_prodotti()

@app.get("/api/vendite")
async def api_get_vendite():
    """Ritorna tutte le vendite"""
    return get_vendite()

@app.post("/api/vendite")
async def api_add_vendita(vendita: Vendita):
    """Aggiunge una vendita"""
    result = add_vendita(vendita)
    return {"success": True, **result}

@app.delete("/api/vendite/{vendita_id}")
async def api_delete_vendita(vendita_id: str):
    """Elimina una vendita"""
    delete_vendita(vendita_id)
    return {"success": True}

@app.get("/api/spese")
async def api_get_spese():
    """Ritorna tutte le spese"""
    return get_spese()

@app.post("/api/spese")
async def api_add_spesa(spesa: Spesa):
    """Aggiunge una spesa"""
    result = add_spesa(spesa)
    return {"success": True, **result}

@app.delete("/api/spese/{spesa_id}")
async def api_delete_spesa(spesa_id: str):
    """Elimina una spesa"""
    delete_spesa(spesa_id)
    return {"success": True}

@app.get("/api/statistiche/{periodo}")
async def api_get_statistiche(periodo: str):
    """Ritorna statistiche per periodo"""
    now = datetime.now()
    
    if periodo == "oggi":
        data_inizio = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif periodo == "settimana":
        data_inizio = now - timedelta(days=7)
    else:  # mese
        data_inizio = now - timedelta(days=30)
    
    stats = calcola_statistiche(data_inizio, now)
    stats["periodo"] = "Oggi" if periodo == "oggi" else "Ultimi 7 giorni" if periodo == "settimana" else "Ultimi 30 giorni"
    
    return stats

@app.post("/api/reset")
async def api_reset(request: ResetRequest):
    """Reset periodo con password"""
    if request.password != CONFIG["password_reset"]:
        raise HTTPException(status_code=403, detail="Password errata")
    
    reset_periodo(request.periodo)
    return {"success": True}

@app.get("/api/status")
async def api_status():
    """Status del sistema"""
    ultimo_report = get_ultimo_report()
    return {
        "online": check_internet(),
        "ultimo_report": ultimo_report.isoformat() if ultimo_report else None,
        "file_dati": str(EXCEL_FILE),
        "file_esiste": EXCEL_FILE.exists()
    }

@app.post("/api/invia-report")
async def api_invia_report():
    """Forza invio report manuale"""
    if not check_internet():
        raise HTTPException(status_code=503, detail="Nessuna connessione internet")
    
    success = invia_email_report()
    if success:
        return {"success": True, "message": "Report inviato!"}
    else:
        raise HTTPException(status_code=500, detail="Errore invio email")

@app.get("/api/download-excel")
async def api_download_excel():
    """Scarica il file Excel dei dati"""
    if not EXCEL_FILE.exists():
        raise HTTPException(status_code=404, detail="File non trovato")
    
    return FileResponse(
        EXCEL_FILE,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"storico_bar_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )

@app.post("/api/import-dati")
async def api_import_dati(dati: dict):
    """Importa dati da vecchia app (vendite e spese)"""
    try:
        vendite_importate = 0
        spese_importate = 0
        
        if "vendite" in dati:
            for v in dati["vendite"]:
                vendita = Vendita(
                    prodotto_id=v.get("prodotto_id", "imported"),
                    nome_prodotto=v.get("nome_prodotto", "Importato"),
                    prezzo=float(v.get("prezzo", 0)),
                    categoria=v.get("categoria", "ALTRO"),
                    timestamp=v.get("timestamp")
                )
                add_vendita(vendita)
                vendite_importate += 1
        
        if "spese" in dati:
            for s in dati["spese"]:
                spesa = Spesa(
                    categoria_spesa=s.get("categoria_spesa", "Importato"),
                    importo=float(s.get("importo", 0)),
                    note=s.get("note", ""),
                    timestamp=s.get("timestamp")
                )
                add_spesa(spesa)
                spese_importate += 1
        
        return {
            "success": True,
            "vendite_importate": vendite_importate,
            "spese_importate": spese_importate
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# ==================== SERVE FRONTEND ====================

# Monta file statici
STATIC_DIR = Path(__file__).parent / "static"
if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

@app.get("/")
async def serve_index():
    """Serve la pagina principale"""
    index_file = Path(__file__).parent / "index.html"
    if index_file.exists():
        return HTMLResponse(index_file.read_text(encoding="utf-8"))
    return HTMLResponse("<h1>Bar Manager</h1><p>index.html non trovato</p>")

# ==================== MAIN ====================

if __name__ == "__main__":
    print(f"""
╔══════════════════════════════════════════╗
║     🍺 BAR MANAGER - RASPBERRY PI 🍺      ║
║         {CONFIG['nome_bar']}          ║
╠══════════════════════════════════════════╣
║  Server: http://0.0.0.0:{CONFIG['porta_server']}             ║
║  Dati: {DATA_DIR}                         
║  Report auto: {'Sì' if CONFIG['report_automatico'] else 'No'}                        
╚══════════════════════════════════════════╝
    """)
    
    # Inizializza Excel
    init_excel()
    
    # Avvia monitor internet in background
    if CONFIG["report_automatico"]:
        monitor_thread = threading.Thread(target=monitor_internet, daemon=True)
        monitor_thread.start()
    
    # Avvia server
    uvicorn.run(app, host="0.0.0.0", port=CONFIG["porta_server"])
