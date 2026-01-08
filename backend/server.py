from fastapi import FastAPI, APIRouter, HTTPException
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
from collections import defaultdict
import io
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI(title="Bar Management API")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# ========== MODELS ==========

class Prodotto(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nome: str
    prezzo: float
    categoria: str
    icona: str = "☕"

class Vendita(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    prodotto_id: str
    nome_prodotto: str
    prezzo: float
    categoria: str
    timestamp: str

class VenditaCreate(BaseModel):
    prodotto_id: str
    prezzo_personalizzato: Optional[float] = None

class Spesa(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    categoria_spesa: str
    importo: float
    note: str = ""
    timestamp: str

class SpesaCreate(BaseModel):
    categoria_spesa: str
    importo: float
    note: str = ""

class StatisticheResponse(BaseModel):
    totale_vendite: int
    totale_incasso: float
    totale_spese: float
    profitto_netto: float
    per_categoria: dict
    prodotti_piu_venduti: List[dict]
    periodo: str

# ========== INIZIALIZZAZIONE PRODOTTI ==========

PRODOTTI_INIZIALI = [
    # CAFFETTERIA
    {"nome": "Caffè", "prezzo": 1.20, "categoria": "CAFFETTERIA", "icona": "☕"},
    {"nome": "Caffè Deca", "prezzo": 1.20, "categoria": "CAFFETTERIA", "icona": "☕"},
    {"nome": "Caffè Corretto", "prezzo": 2.00, "categoria": "CAFFETTERIA", "icona": "🥃"},
    
    # BEVANDE
    {"nome": "Caraffa di Vino 0,5L", "prezzo": 6.00, "categoria": "BEVANDE", "icona": "🍷"},
    {"nome": "Caraffa di Vino 1LT", "prezzo": 11.00, "categoria": "BEVANDE", "icona": "🍷"},
    {"nome": "Calice di Vino", "prezzo": 1.50, "categoria": "BEVANDE", "icona": "🍷"},
    {"nome": "Liquori, Grappe, Vodka e Amari", "prezzo": 2.50, "categoria": "BEVANDE", "icona": "🥃"},
    {"nome": "Bibite in Lattina", "prezzo": 2.20, "categoria": "BEVANDE", "icona": "🥤"},
    {"nome": "Bottiglietta d'Acqua", "prezzo": 1.00, "categoria": "BEVANDE", "icona": "💧"},
    
    # GELATI
    {"nome": "Cremino", "prezzo": 1.20, "categoria": "GELATI", "icona": "🍦"},
    {"nome": "Cucciolone", "prezzo": 1.50, "categoria": "GELATI", "icona": "🍦"},
    {"nome": "Magnum, Soia e altri Gelati", "prezzo": 2.00, "categoria": "GELATI", "icona": "🍦"},
    
    # PERSONALIZZATE
    {"nome": "Bigliardo", "prezzo": 0.00, "categoria": "PERSONALIZZATE", "icona": "🎱"},
    {"nome": "Extra", "prezzo": 0.00, "categoria": "PERSONALIZZATE", "icona": "➕"},
]

@app.on_event("startup")
async def startup_event():
    """Inizializza i prodotti nel database se non esistono"""
    count = await db.prodotti.count_documents({})
    if count == 0:
        prodotti = []
        for p in PRODOTTI_INIZIALI:
            prod = Prodotto(**p)
            prodotti.append(prod.model_dump())
        await db.prodotti.insert_many(prodotti)
        logging.info(f"Inizializzati {len(prodotti)} prodotti nel database")

# ========== API ENDPOINTS ==========

@api_router.get("/")
async def root():
    return {"message": "Bar Management API - Proloco Santa Bianca"}

@api_router.get("/prodotti", response_model=List[Prodotto])
async def get_prodotti():
    """Ottieni tutti i prodotti organizzati per categoria"""
    prodotti = await db.prodotti.find({}, {"_id": 0}).to_list(1000)
    # Ordina per categoria e nome
    categorie_ordine = {"CAFFETTERIA": 0, "BEVANDE": 1, "GELATI": 2}
    prodotti.sort(key=lambda x: (categorie_ordine.get(x['categoria'], 999), x['nome']))
    return prodotti

@api_router.post("/vendite", response_model=Vendita)
async def crea_vendita(vendita_input: VenditaCreate):
    """Registra una nuova vendita"""
    # Trova il prodotto
    prodotto = await db.prodotti.find_one({"id": vendita_input.prodotto_id}, {"_id": 0})
    if not prodotto:
        raise HTTPException(status_code=404, detail="Prodotto non trovato")
    
    # Usa prezzo personalizzato se fornito, altrimenti usa prezzo del prodotto
    prezzo_finale = vendita_input.prezzo_personalizzato if vendita_input.prezzo_personalizzato is not None else prodotto['prezzo']
    
    # Validazione prezzo
    if prezzo_finale < 0:
        raise HTTPException(status_code=400, detail="Prezzo non valido")
    
    # Crea la vendita
    vendita = Vendita(
        prodotto_id=prodotto['id'],
        nome_prodotto=prodotto['nome'],
        prezzo=prezzo_finale,
        categoria=prodotto['categoria'],
        timestamp=datetime.now(timezone.utc).isoformat()
    )
    
    await db.vendite.insert_one(vendita.model_dump())
    return vendita

@api_router.get("/vendite", response_model=List[Vendita])
async def get_vendite(
    data_inizio: Optional[str] = None,
    data_fine: Optional[str] = None,
    limite: int = 1000
):
    """Ottieni tutte le vendite con filtri opzionali"""
    query = {}
    
    if data_inizio or data_fine:
        query['timestamp'] = {}
        if data_inizio:
            query['timestamp']['$gte'] = data_inizio
        if data_fine:
            query['timestamp']['$lte'] = data_fine
    
    vendite = await db.vendite.find(query, {"_id": 0}).sort("timestamp", -1).to_list(limite)
    return vendite

@api_router.get("/statistiche/oggi", response_model=StatisticheResponse)
async def statistiche_oggi():
    """Statistiche delle vendite di oggi"""
    oggi = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    domani = oggi + timedelta(days=1)
    
    return await calcola_statistiche(
        oggi.isoformat(),
        domani.isoformat(),
        "Oggi"
    )

@api_router.get("/statistiche/settimana", response_model=StatisticheResponse)
async def statistiche_settimana():
    """Statistiche delle vendite della settimana corrente (ultimi 7 giorni)"""
    oggi = datetime.now(timezone.utc)
    sette_giorni_fa = oggi - timedelta(days=7)
    
    return await calcola_statistiche(
        sette_giorni_fa.isoformat(),
        oggi.isoformat(),
        "Ultimi 7 giorni"
    )

@api_router.get("/statistiche/mese", response_model=StatisticheResponse)
async def statistiche_mese():
    """Statistiche delle vendite del mese corrente (ultimi 30 giorni)"""
    oggi = datetime.now(timezone.utc)
    trenta_giorni_fa = oggi - timedelta(days=30)
    
    return await calcola_statistiche(
        trenta_giorni_fa.isoformat(),
        oggi.isoformat(),
        "Ultimi 30 giorni"
    )

@api_router.delete("/vendite/{vendita_id}")
async def elimina_vendita(vendita_id: str):
    """Elimina una vendita (per correzioni)"""
    result = await db.vendite.delete_one({"id": vendita_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Vendita non trovata")
    return {"message": "Vendita eliminata con successo"}

# ========== SPESE ENDPOINTS ==========

@api_router.post("/spese", response_model=Spesa)
async def crea_spesa(spesa_input: SpesaCreate):
    """Registra una nuova spesa"""
    if spesa_input.importo <= 0:
        raise HTTPException(status_code=400, detail="Importo non valido")
    
    spesa = Spesa(
        categoria_spesa=spesa_input.categoria_spesa,
        importo=spesa_input.importo,
        note=spesa_input.note,
        timestamp=datetime.now(timezone.utc).isoformat()
    )
    
    await db.spese.insert_one(spesa.model_dump())
    return spesa

@api_router.get("/spese", response_model=List[Spesa])
async def get_spese(
    data_inizio: Optional[str] = None,
    data_fine: Optional[str] = None,
    limite: int = 1000
):
    """Ottieni tutte le spese con filtri opzionali"""
    query = {}
    
    if data_inizio or data_fine:
        query['timestamp'] = {}
        if data_inizio:
            query['timestamp']['$gte'] = data_inizio
        if data_fine:
            query['timestamp']['$lte'] = data_fine
    
    spese = await db.spese.find(query, {"_id": 0}).sort("timestamp", -1).to_list(limite)
    return spese

@api_router.delete("/spese/{spesa_id}")
async def elimina_spesa(spesa_id: str):
    """Elimina una spesa (per correzioni)"""
    result = await db.spese.delete_one({"id": spesa_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Spesa non trovata")
    return {"message": "Spesa eliminata con successo"}

@api_router.get("/export/csv")
async def export_csv(
    data_inizio: Optional[str] = None,
    data_fine: Optional[str] = None
):
    """Esporta le vendite in formato CSV"""
    vendite = await get_vendite(data_inizio, data_fine, limite=10000)
    
    # Crea CSV
    output = io.StringIO()
    output.write("Data,Ora,Prodotto,Categoria,Prezzo\n")
    
    for v in vendite:
        dt = datetime.fromisoformat(v.timestamp)
        data = dt.strftime("%d/%m/%Y")
        ora = dt.strftime("%H:%M:%S")
        output.write(f'{data},{ora},{v.nome_prodotto},{v.categoria},{v.prezzo:.2f}\n')
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=vendite.csv"}
    )

@api_router.get("/export/excel-completo")
async def export_excel_completo():
    """Esporta tutto lo storico in Excel suddiviso per mesi e settimane"""
    # Ottieni tutte le vendite e spese
    vendite = await db.vendite.find({}, {"_id": 0}).sort("timestamp", 1).to_list(100000)
    spese = await db.spese.find({}, {"_id": 0}).sort("timestamp", 1).to_list(100000)
    
    # Crea workbook
    wb = Workbook()
    wb.remove(wb.active)  # Rimuovi foglio di default
    
    # Raggruppa per mese
    vendite_per_mese = defaultdict(list)
    spese_per_mese = defaultdict(list)
    
    for v in vendite:
        dt = datetime.fromisoformat(v['timestamp'])
        mese_key = dt.strftime("%Y-%m")
        vendite_per_mese[mese_key].append(v)
    
    for s in spese:
        dt = datetime.fromisoformat(s['timestamp'])
        mese_key = dt.strftime("%Y-%m")
        spese_per_mese[mese_key].append(s)
    
    # Unisci tutti i mesi
    tutti_mesi = sorted(set(list(vendite_per_mese.keys()) + list(spese_per_mese.keys())), reverse=True)
    
    if not tutti_mesi:
        # Crea foglio vuoto se non ci sono dati
        ws = wb.create_sheet("Nessun Dato")
        ws['A1'] = "Nessuna transazione registrata"
        ws['A1'].font = Font(size=14, bold=True)
    else:
        # Crea un foglio per ogni mese
        for mese in tutti_mesi[:12]:  # Limita a ultimi 12 mesi
            dt_mese = datetime.strptime(mese, "%Y-%m")
            nome_mese = dt_mese.strftime("%B %Y")
            
            ws = wb.create_sheet(nome_mese)
            
            # Header styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            
            # Sezione VENDITE
            ws['A1'] = f"VENDITE - {nome_mese}"
            ws['A1'].font = Font(size=14, bold=True, color="00B050")
            ws.merge_cells('A1:E1')
            
            # Headers vendite
            headers_vendite = ['Data', 'Ora', 'Prodotto', 'Categoria', 'Importo €']
            for col, header in enumerate(headers_vendite, 1):
                cell = ws.cell(row=2, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Dati vendite
            row = 3
            totale_vendite = 0
            for v in vendite_per_mese.get(mese, []):
                dt = datetime.fromisoformat(v['timestamp'])
                ws.cell(row=row, column=1, value=dt.strftime("%d/%m/%Y"))
                ws.cell(row=row, column=2, value=dt.strftime("%H:%M:%S"))
                ws.cell(row=row, column=3, value=v['nome_prodotto'])
                ws.cell(row=row, column=4, value=v['categoria'])
                ws.cell(row=row, column=5, value=v['prezzo'])
                totale_vendite += v['prezzo']
                row += 1
            
            # Totale vendite
            ws.cell(row=row, column=4, value="TOTALE VENDITE:").font = Font(bold=True)
            ws.cell(row=row, column=5, value=totale_vendite).font = Font(bold=True, color="00B050")
            
            # Sezione SPESE (lascia 2 righe vuote)
            row += 3
            ws.cell(row=row, column=1, value=f"SPESE - {nome_mese}").font = Font(size=14, bold=True, color="FF0000")
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            # Headers spese
            headers_spese = ['Data', 'Ora', 'Categoria Spesa', 'Note', 'Importo €']
            for col, header in enumerate(headers_spese, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal='center')
            row += 1
            
            # Dati spese
            totale_spese = 0
            for s in spese_per_mese.get(mese, []):
                dt = datetime.fromisoformat(s['timestamp'])
                ws.cell(row=row, column=1, value=dt.strftime("%d/%m/%Y"))
                ws.cell(row=row, column=2, value=dt.strftime("%H:%M:%S"))
                ws.cell(row=row, column=3, value=s['categoria_spesa'])
                ws.cell(row=row, column=4, value=s.get('note', ''))
                ws.cell(row=row, column=5, value=s['importo'])
                totale_spese += s['importo']
                row += 1
            
            # Totale spese
            ws.cell(row=row, column=4, value="TOTALE SPESE:").font = Font(bold=True)
            ws.cell(row=row, column=5, value=totale_spese).font = Font(bold=True, color="C00000")
            
            # RIEPILOGO MESE
            row += 3
            ws.cell(row=row, column=1, value="RIEPILOGO MESE").font = Font(size=14, bold=True)
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            ws.cell(row=row, column=3, value="Incassi Totali:")
            ws.cell(row=row, column=4, value=totale_vendite).font = Font(color="00B050", bold=True)
            row += 1
            
            ws.cell(row=row, column=3, value="Spese Totali:")
            ws.cell(row=row, column=4, value=totale_spese).font = Font(color="C00000", bold=True)
            row += 1
            
            profitto = totale_vendite - totale_spese
            ws.cell(row=row, column=3, value="PROFITTO NETTO:").font = Font(bold=True, size=12)
            cell_profitto = ws.cell(row=row, column=4, value=profitto)
            cell_profitto.font = Font(bold=True, size=12, color="00B050" if profitto >= 0 else "C00000")
            
            # Larghezza colonne
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 12
    
    # Salva in memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=storico_completo.xlsx"}
    )

@api_router.post("/reset-periodo")
async def reset_periodo(password: str, periodo: str):
    """Reset vendite e spese per un periodo specifico (richiede password)"""
    if password != "5054":
        raise HTTPException(status_code=403, detail="Password errata")
    
    # Calcola date in base al periodo
    oggi = datetime.now(timezone.utc)
    
    if periodo == "oggi":
        data_inizio = oggi.replace(hour=0, minute=0, second=0, microsecond=0)
        data_fine = oggi
    elif periodo == "settimana":
        data_inizio = oggi - timedelta(days=7)
        data_fine = oggi
    elif periodo == "mese":
        data_inizio = oggi - timedelta(days=30)
        data_fine = oggi
    else:
        raise HTTPException(status_code=400, detail="Periodo non valido")
    
    # Elimina vendite e spese del periodo
    result_vendite = await db.vendite.delete_many({
        "timestamp": {
            "$gte": data_inizio.isoformat(),
            "$lte": data_fine.isoformat()
        }
    })
    
    result_spese = await db.spese.delete_many({
        "timestamp": {
            "$gte": data_inizio.isoformat(),
            "$lte": data_fine.isoformat()
        }
    })
    
    return {
        "message": f"Reset {periodo} completato",
        "vendite_eliminate": result_vendite.deleted_count,
        "spese_eliminate": result_spese.deleted_count
    }

# ========== FUNZIONI HELPER ==========

async def calcola_statistiche(data_inizio: str, data_fine: str, periodo: str) -> StatisticheResponse:
    """Calcola statistiche per un periodo specifico"""
    vendite = await db.vendite.find({
        "timestamp": {"$gte": data_inizio, "$lte": data_fine}
    }, {"_id": 0}).to_list(10000)
    
    spese = await db.spese.find({
        "timestamp": {"$gte": data_inizio, "$lte": data_fine}
    }, {"_id": 0}).to_list(10000)
    
    if not vendite:
        totale_spese = sum(s['importo'] for s in spese) if spese else 0.0
        return StatisticheResponse(
            totale_vendite=0,
            totale_incasso=0.0,
            totale_spese=round(totale_spese, 2),
            profitto_netto=round(-totale_spese, 2),
            per_categoria={},
            prodotti_piu_venduti=[],
            periodo=periodo
        )
    
    # Calcola totali
    totale_vendite = len(vendite)
    totale_incasso = sum(v['prezzo'] for v in vendite)
    totale_spese = sum(s['importo'] for s in spese) if spese else 0.0
    profitto_netto = totale_incasso - totale_spese
    
    # Raggruppa per categoria
    per_categoria = defaultdict(lambda: {"vendite": 0, "incasso": 0.0})
    for v in vendite:
        cat = v['categoria']
        per_categoria[cat]['vendite'] += 1
        per_categoria[cat]['incasso'] += v['prezzo']
    
    # Conta prodotti più venduti
    prodotti_count = defaultdict(lambda: {"nome": "", "count": 0, "incasso": 0.0})
    for v in vendite:
        pid = v['prodotto_id']
        prodotti_count[pid]['nome'] = v['nome_prodotto']
        prodotti_count[pid]['count'] += 1
        prodotti_count[pid]['incasso'] += v['prezzo']
    
    # Top 5 prodotti più venduti
    prodotti_piu_venduti = sorted(
        [{
            "nome": data['nome'],
            "quantita": data['count'],
            "incasso": round(data['incasso'], 2)
        } for data in prodotti_count.values()],
        key=lambda x: x['quantita'],
        reverse=True
    )[:5]
    
    return StatisticheResponse(
        totale_vendite=totale_vendite,
        totale_incasso=round(totale_incasso, 2),
        totale_spese=round(totale_spese, 2),
        profitto_netto=round(profitto_netto, 2),
        per_categoria=dict(per_categoria),
        prodotti_piu_venduti=prodotti_piu_venduti,
        periodo=periodo
    )

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
